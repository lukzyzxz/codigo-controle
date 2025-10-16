import pandas as pd
import os
import pwinput
import time
import shutil
from datetime import datetime, timedelta
from tabulate import tabulate
from pathlib import Path
from colorama import Fore, Style, init
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

init(autoreset=True)

# Configurações do sistema
ARQ_ALUNOS = Path("alunos.csv")
ARQ_ALUNOS_XLSX = Path("alunos.xlsx")
ARQ_REL = Path("relatorios.csv")
ARQ_REL_XLSX = Path("relatorios.xlsx")
ARQ_AG = Path("agendamentos.csv")
USERS = {
    "admin": {"senha": "admin123", "nome": "Administrador"},
    "proftec": {"senha": "tecnico123", "nome": "Prof. Técnico"},
    "professor": {"senha": "prof123", "nome": "Professor"}
}

class AgendamentoService:
    """Serviço para gerenciamento de agendamentos"""
    
    def __init__(self):
        self.arquivo_ag = ARQ_AG
        self.arquivo_xlsx = Path("agendamentos.xlsx")
    
    def carregar_agendamentos(self) -> pd.DataFrame:
        """Carrega os agendamentos do arquivo"""
        if not self.arquivo_ag.exists():
            self._criar_agendamentos_iniciais()
        return carregar_dataframe(self.arquivo_ag)
    
    def _criar_agendamentos_iniciais(self):
        """Cria a estrutura inicial de agendamentos"""
        horarios = [f"{h:02d}:00 - {h+1:02d}:00" for h in range(8, 21)]
        pcs = [f"PC{i:02d}" for i in range(1, 21)]
        registros = []
        for pc in pcs:
            for h in horarios:
                registros.append([pc, h, "livre", "Disponível"])
        df_ag = pd.DataFrame(registros, columns=["pc", "horario", "professor", "status"])
        salvar_csv_xlsx(df_ag, self.arquivo_ag, self.arquivo_xlsx)
    
    def verificar_conflito(self, professor: str, horario: str, pc: str = None) -> bool:
        """Verifica se professor já tem agendamento no mesmo horário para o mesmo PC"""
        agendamentos = self.carregar_agendamentos()
        
        if pc:
            # Verifica conflito específico para um PC
            conflito = agendamentos[
                (agendamentos["professor"] == professor) &
                (agendamentos["horario"] == horario) &
                (agendamentos["pc"] == pc) &
                (agendamentos["status"] == "Agendado")
            ]
        else:
            # Verifica se o professor já tem qualquer agendamento neste horário
            conflito = agendamentos[
                (agendamentos["professor"] == professor) &
                (agendamentos["horario"] == horario) &
                (agendamentos["status"] == "Agendado")
            ]
        
        return not conflito.empty
    
    def agendar_horario(self, idx: int, professor: str) -> bool:
        """Realiza o agendamento de um horário"""
        try:
            df_agend = self.carregar_agendamentos()
            
            if idx not in df_agend.index:
                return False
            
            horario_escolhido = df_agend.loc[idx, "horario"]
            pc_escolhido = df_agend.loc[idx, "pc"]
            
            # Verificar conflito apenas para o PC específico
            if self.verificar_conflito(professor, horario_escolhido, pc_escolhido):
                return False
            
            # Realizar agendamento
            df_agend.loc[idx, "professor"] = professor
            df_agend.loc[idx, "status"] = "Agendado"
            salvar_csv_xlsx(df_agend, self.arquivo_ag, self.arquivo_xlsx)
            return True
            
        except Exception as e:
            msg(f"Erro ao agendar: {e}", "err")
            return False
    
    def agendar_multiplos_pcs(self, indices: list, professor: str) -> dict:
        """Agenda múltiplos PCs no mesmo horário"""
        resultados = {
            'sucessos': [],
            'falhas': []
        }
        
        try:
            df_agend = self.carregar_agendamentos()
            horarios_agendados = set()
            pcs_agendados = set()
            
            # Primeiro, verificar todos os agendamentos
            for idx in indices:
                if idx not in df_agend.index:
                    resultados['falhas'].append(f"Índice {idx} inválido")
                    continue
                
                horario = df_agend.loc[idx, "horario"]
                pc = df_agend.loc[idx, "pc"]
                
                # Verificar se o horário é o mesmo para todos
                horarios_agendados.add(horario)
                pcs_agendados.add(pc)
                
                # Verificar conflito para este PC específico
                if self.verificar_conflito(professor, horario, pc):
                    resultados['falhas'].append(f"PC {pc} já agendado por você neste horário")
                    continue
            
            # Se há mais de um horário diferente, não permitir
            if len(horarios_agendados) > 1:
                resultados['falhas'].append("Todos os agendamentos devem ser no mesmo horário")
                return resultados
            
            # Realizar os agendamentos válidos
            for idx in indices:
                if idx in df_agend.index:
                    horario = df_agend.loc[idx, "horario"]
                    pc = df_agend.loc[idx, "pc"]
                    
                    if not self.verificar_conflito(professor, horario, pc):
                        df_agend.loc[idx, "professor"] = professor
                        df_agend.loc[idx, "status"] = "Agendado"
                        resultados['sucessos'].append(f"PC {pc} - {horario}")
            
            if resultados['sucessos']:
                salvar_csv_xlsx(df_agend, self.arquivo_ag, self.arquivo_xlsx)
            
            return resultados
            
        except Exception as e:
            msg(f"Erro ao agendar múltiplos PCs: {e}", "err")
            resultados['falhas'].append("Erro interno do sistema")
            return resultados
    
    def get_horarios_disponiveis(self) -> pd.DataFrame:
        """Retorna horários disponíveis"""
        df = self.carregar_agendamentos()
        return df[df["status"] == "Disponível"]
    
    def get_horarios_agendados(self) -> pd.DataFrame:
        """Retorna horários agendados"""
        df = self.carregar_agendamentos()
        return df[df["status"] == "Agendado"]
    
    def get_horarios_agrupados(self) -> dict:
        """Retorna horários agrupados por período"""
        df = self.carregar_agendamentos()
        horarios_agrupados = {}
        
        for horario in df['horario'].unique():
            pcs_horario = df[df['horario'] == horario]
            horarios_agrupados[horario] = pcs_horario
        
        return horarios_agrupados

def criar_backup():
    """Cria backup dos arquivos importantes"""
    try:
        backup_dir = Path("backup") / datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir.mkdir(parents=True, exist_ok=True)
        
        arquivos = [ARQ_ALUNOS, ARQ_REL, ARQ_AG, ARQ_ALUNOS_XLSX, ARQ_REL_XLSX]
        for arquivo in arquivos:
            if arquivo.exists():
                shutil.copy2(arquivo, backup_dir / arquivo.name)
        
        msg(f"Backup criado em: {backup_dir}", "ok")
        return True
    except Exception as e:
        msg(f"Erro ao criar backup: {e}", "err")
        return False

def validar_pc_existente(numero_pc: str) -> bool:
    """Valida se o PC existe no laboratório"""
    try:
        pcs_validos = [f"PC{i:02d}" for i in range(1, 21)]
        pc = f"PC{numero_pc.zfill(2)}"
        return pc in pcs_validos
    except Exception:
        return False

def validar_duracao_minima(entrada: str, saida: str) -> tuple[bool, str]:
    """Valida tempo mínimo de uso (30 minutos)"""
    try:
        duracao = calcular_duracao("01/01/2000", entrada, saida)
        horas, minutos = map(int, duracao.split(':'))
        tempo_total_minutos = horas * 60 + minutos
        if tempo_total_minutos < 30:
            return False, "Tempo mínimo de uso é 30 minutos"
        return True, "Duração válida"
    except Exception:
        return False, "Erro ao validar duração"

def validar_hora_agendamento(hora_str: str, data_str: str = None) -> tuple[bool, str]:
    """Valida se o horário está dentro do expediente e não é no passado"""
    try:
        hora = datetime.strptime(hora_str, "%H:%M").time()
        hora_min = datetime.strptime("08:00", "%H:%M").time()
        hora_max = datetime.strptime("21:00", "%H:%M").time()
        
        if not (hora_min <= hora <= hora_max):
            return False, "Horário fora do expediente (08:00 - 21:00)"
            
        if data_str:
            data_hora = datetime.strptime(f"{data_str} {hora_str}", "%d/%m/%Y %H:%M")
            if data_hora < datetime.now():
                return False, "Não é possível agendar para horários no passado"
                
        return True, "Horário válido"
    except ValueError:
        return False, "Formato de hora inválido (use HH:MM)"

def limpar_tela():
    os.system("cls" if os.name == "nt" else "clear")

def msg(text, tipo="info"):
    cores = {"info": Fore.CYAN, "ok": Fore.GREEN, "warn": Fore.YELLOW, "err": Fore.RED}
    print(cores.get(tipo, Fore.CYAN) + text + Style.RESET_ALL)

def pedir_validado(prompt, func):
    while True:
        v = input(prompt).strip()
        r = func(v)
        if r is not None:
            return r

def validar_nome(nome: str):
    if nome.replace(" ", "").isalpha():
        return nome.title().strip()
    return None

def validar_numero(texto: str):
    if texto.isdigit():
        return texto
    return None

def validar_data(data_str: str):
    try:
        d = datetime.strptime(data_str, "%d/%m/%Y")
        return d.strftime("%d/%m/%Y")
    except Exception:
        return None

def validar_hora(hora_str: str):
    try:
        h = datetime.strptime(hora_str, "%H:%M")
        return h.strftime("%H:%M")
    except Exception:
        return None

def confirmar_sn(mensagem: str):
    while True:
        r = input(mensagem + " (s/n): ").lower().strip()
        if r in ("s", "n"):
            return r

def salvar_csv_xlsx(df: pd.DataFrame, csv_path: Path, xlsx_path: Path):
    df.to_csv(csv_path, index=False)
    try:
        df.to_excel(xlsx_path, index=False)
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            header_font = Font(bold=True)
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            wb.save(xlsx_path)
        except Exception:
            pass
    except Exception:
        df.to_csv(csv_path, index=False)

def calcular_duracao(data_str: str, entrada_str: str, saida_str: str) -> str:
    try:
        data = datetime.strptime(data_str, "%d/%m/%Y")
        entrada = datetime.strptime(entrada_str, "%H:%M").time()
        saida = datetime.strptime(saida_str, "%H:%M").time()
        dt_entrada = datetime.combine(data.date(), entrada)
        dt_saida = datetime.combine(data.date(), saida)
        if dt_saida < dt_entrada:
            dt_saida += timedelta(days=1)
        delta = dt_saida - dt_entrada
        horas = int(delta.total_seconds() // 3600)
        minutos = int((delta.total_seconds() % 3600) // 60)
        return f"{horas:02d}:{minutos:02d}"
    except Exception:
        return ""

def carregar_dataframe(path: Path, cols=None):
    if not path.exists():
        if cols:
            return pd.DataFrame(columns=cols)
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except Exception:
        return pd.DataFrame()

def menu_computadores(usuario_logado: str):
    while True:
        print("\n=== MENU COMPUTADORES ===")
        print("1 - Registrar novo aluno")
        print("2 - Consultar alunos cadastrados")
        print("3 - Editar aluno")
        print("4 - Excluir aluno")
        print("5 - Voltar ao menu principal")
        escolha = input("Escolha uma opção: ").strip()

        if escolha == "1":
            numero_pc = pedir_validado("Digite o número do PC (ex: 01, 02...): ", validar_numero)
            
            if not validar_pc_existente(numero_pc):
                msg("Número de PC inválido. Use números de 01 a 20.", "err")
                continue
                
            pc = f"PC{numero_pc.zfill(2)}"
            nome = pedir_validado("Digite o nome do aluno: ", validar_nome)

            print("\nDeseja usar a data e hora atuais para o registro?")
            print("1 - Sim, usar data e hora atuais")
            print("2 - Não, quero inserir manualmente")
            opc = ""
            while opc not in ("1", "2"):
                opc = input("Escolha uma opção (1/2): ").strip()

            if opc == "1":
                agora = datetime.now()
                data_automatica = agora.strftime("%d/%m/%Y")
                hora_automatica = agora.strftime("%H:%M")
                print(f"\nData atual: {data_automatica}")
                print(f"Horário atual: {hora_automatica}")
                if confirmar_sn("Deseja confirmar essa data e hora?") == "s":
                    data = data_automatica
                    entrada = hora_automatica
                    msg("Data e hora registradas automaticamente.", "ok")
                else:
                    data = None
                    entrada = None
            else:
                data = None
                entrada = None

            if not data:
                data = pedir_validado("Digite a data (DD/MM/AAAA): ", validar_data)
            if not entrada:
                while True:
                    entrada = pedir_validado("Digite o horário de entrada (HH:MM): ", validar_hora)
                    valido, mensagem = validar_hora_agendamento(entrada, data)
                    if valido:
                        break
                    msg(mensagem, "warn")
                    
            while True:
                saida = pedir_validado("Digite o horário de saída (HH:MM): ", validar_hora)
                valido, mensagem = validar_hora_agendamento(saida, data)
                if valido:
                    break
                msg(mensagem, "warn")

            duracao_valida, mensagem_duracao = validar_duracao_minima(entrada, saida)
            if not duracao_valida:
                msg(mensagem_duracao, "warn")
                if confirmar_sn("Deseja continuar mesmo assim?") == "n":
                    continue

            duracao = calcular_duracao(data, entrada, saida)

            novo_registro = pd.DataFrame([{
                "pc": pc,
                "nome": nome,
                "data": data,
                "entrada": entrada,
                "saida": saida,
                "duracao": duracao
            }])

            df = carregar_dataframe(ARQ_ALUNOS, cols=["pc", "nome", "data", "entrada", "saida", "duracao"])
            if not df.empty:
                df = pd.concat([df, novo_registro], ignore_index=True)
            else:
                df = novo_registro

            salvar_csv_xlsx(df, ARQ_ALUNOS, ARQ_ALUNOS_XLSX)
            criar_backup()
            msg("Registro salvo com sucesso!", "ok")

        elif escolha == "2":
            msg("Carregando dados dos alunos...", "info")
            time.sleep(0.6)
            dados = carregar_dataframe(ARQ_ALUNOS)
            if dados.empty:
                msg("Nenhum aluno cadastrado ainda.", "warn")
            else:
                print("\n=== Alunos Cadastrados ===")
                print(tabulate(dados, headers="keys", tablefmt="grid", showindex=True))

        elif escolha == "3":
            if not ARQ_ALUNOS.exists():
                msg("Nenhum arquivo encontrado para edição.", "warn")
                continue
            dados = carregar_dataframe(ARQ_ALUNOS)
            if dados.empty:
                msg("Nenhum aluno cadastrado para editar.", "warn")
                continue
            print("\nAlunos cadastrados:")
            print(tabulate(dados, headers="keys", tablefmt="grid", showindex=True))
            try:
                idx = int(input("Digite o índice do aluno que deseja editar: ").strip())
            except Exception:
                msg("Entrada inválida.", "warn")
                continue
            if idx not in dados.index:
                msg("Índice inválido.", "warn")
                continue
            print("\nDeixe em branco para não alterar.")
            novo_pc = input(f"PC atual ({dados.loc[idx,'pc']}): ").strip()
            if novo_pc:
                num = validar_numero(novo_pc.replace("PC", "").replace("pc", ""))
                if num and validar_pc_existente(num):
                    novo_pc = f"PC{num.zfill(2)}"
                else:
                    msg("Número de PC inválido. Mantendo valor anterior.", "warn")
                    novo_pc = dados.loc[idx, "pc"]
            else:
                novo_pc = dados.loc[idx, "pc"]

            novo_nome = input(f"Nome atual ({dados.loc[idx,'nome']}): ").strip()
            if novo_nome:
                valid = validar_nome(novo_nome)
                novo_nome = valid if valid else dados.loc[idx, "nome"]
            else:
                novo_nome = dados.loc[idx, "nome"]

            nova_data = input(f"Data atual ({dados.loc[idx,'data']}): ").strip()
            if nova_data:
                nova_data = validar_data(nova_data) or dados.loc[idx, "data"]
            else:
                nova_data = dados.loc[idx, "data"]

            nova_entrada = input(f"Entrada atual ({dados.loc[idx,'entrada']}): ").strip()
            if nova_entrada:
                nova_entrada = validar_hora(nova_entrada) or dados.loc[idx, "entrada"]
            else:
                nova_entrada = dados.loc[idx, "entrada"]

            nova_saida = input(f"Saída atual ({dados.loc[idx,'saida']}): ").strip()
            if nova_saida:
                nova_saida = validar_hora(nova_saida) or dados.loc[idx, "saida"]
            else:
                nova_saida = dados.loc[idx, "saida"]

            duracao = calcular_duracao(nova_data, nova_entrada, nova_saida)

            dados.loc[idx, "pc"] = novo_pc
            dados.loc[idx, "nome"] = novo_nome
            dados.loc[idx, "data"] = nova_data
            dados.loc[idx, "entrada"] = nova_entrada
            dados.loc[idx, "saida"] = nova_saida
            dados.loc[idx, "duracao"] = duracao

            salvar_csv_xlsx(dados, ARQ_ALUNOS, ARQ_ALUNOS_XLSX)
            criar_backup()
            msg("Registro atualizado com sucesso!", "ok")

        elif escolha == "4":
            if not ARQ_ALUNOS.exists():
                msg("Nenhum arquivo encontrado para exclusão.", "warn")
                continue
            dados = carregar_dataframe(ARQ_ALUNOS)
            if dados.empty:
                msg("Nenhum aluno cadastrado para excluir.", "warn")
                continue
            print("\nAlunos cadastrados:")
            print(tabulate(dados, headers="keys", tablefmt="grid", showindex=True))
            try:
                idx = int(input("Digite o índice do aluno que deseja excluir: ").strip())
            except Exception:
                msg("Entrada inválida.", "warn")
                continue
            if idx not in dados.index:
                msg("Índice inválido.", "warn")
                continue
            if confirmar_sn(f"Tem certeza que deseja excluir o registro de {dados.loc[idx,'nome']} no dia {dados.loc[idx,'data']}?") == "s":
                msg("Apagando registro...", "info")
                time.sleep(0.6)
                dados = dados.drop(idx).reset_index(drop=True)
                salvar_csv_xlsx(dados, ARQ_ALUNOS, ARQ_ALUNOS_XLSX)
                criar_backup()
                msg("Registro excluído com sucesso!", "ok")
            else:
                msg("Exclusão cancelada.", "warn")

        elif escolha == "5":
            return
        else:
            msg("Opção inválida.", "warn")

def gerar_relatorio(usuario_logado: str):
    print("\n=== RELATÓRIO DE AULA ===")
    professor = pedir_validado("Digite seu nome (professor): ", validar_nome)
    descricao = input("Digite o relatório da aula: ").strip()
    msg("Salvando relatório...", "info")
    time.sleep(0.6)
    novo_rel = pd.DataFrame([{"professor": professor, "relatorio": descricao, "usuario": usuario_logado}])
    df = carregar_dataframe(ARQ_REL, cols=["professor", "relatorio", "usuario"])
    if not df.empty:
        df = pd.concat([df, novo_rel], ignore_index=True)
    else:
        df = novo_rel
    salvar_csv_xlsx(df, ARQ_REL, ARQ_REL_XLSX)
    criar_backup()
    msg("Relatório salvo com sucesso!", "ok")

def menu_agendamento(usuario_logado: str):
    agendamento_service = AgendamentoService()
    
    while True:
        print("\n=== AGENDAMENTOS ===")
        print("1 - Ver horários disponíveis")
        print("2 - Ver horários já agendados")
        print("3 - Agendar horário (PC único)")
        print("4 - Agendar múltiplos PCs")
        print("5 - Voltar ao menu principal")
        escolha = input("Escolha uma opção: ").strip()

        if escolha == "1":
            disponiveis = agendamento_service.get_horarios_disponiveis()
            if disponiveis.empty:
                msg("Não há horários disponíveis", "warn")
            else:
                print("\nHorários disponíveis:")
                print(tabulate(disponiveis[["pc", "horario"]], headers="keys", tablefmt="grid", showindex=True))

        elif escolha == "2":
            agendados = agendamento_service.get_horarios_agendados()
            if agendados.empty:
                msg("Nenhum horário agendado", "warn")
            else:
                print("\nHorários agendados:")
                print(tabulate(agendados[["pc", "horario", "professor"]], headers="keys", tablefmt="grid", showindex=True))

        elif escolha == "3":
            disponiveis = agendamento_service.get_horarios_disponiveis()
            if disponiveis.empty:
                msg("Nenhum horário disponível para agendamento", "warn")
                return
            print("\nHorários disponíveis:")
            print(tabulate(disponiveis[["pc", "horario"]], headers="keys", tablefmt="grid", showindex=True))
            try:
                escolha_idx = int(input("Digite o número do horário que deseja agendar: ").strip())
            except Exception:
                msg("Entrada inválida.", "warn")
                return
                
            if agendamento_service.agendar_horario(escolha_idx, USERS[usuario_logado]["nome"]):
                criar_backup()
                msg("Agendamento realizado com sucesso!", "ok")
            else:
                msg("Não foi possível realizar o agendamento. Verifique se você já tem um horário neste período.", "err")

        elif escolha == "4":
            disponiveis = agendamento_service.get_horarios_disponiveis()
            if disponiveis.empty:
                msg("Nenhum horário disponível para agendamento", "warn")
                continue

            # Agrupar horários disponíveis
            horarios_agrupados = agendamento_service.get_horarios_agrupados()
            
            print("\nHorários disponíveis agrupados:")
            for horario, pcs in horarios_agrupados.items():
                pcs_disponiveis = pcs[pcs['status'] == 'Disponível']
                if not pcs_disponiveis.empty:
                    print(f"\n🕒 {horario}:")
                    print(tabulate(pcs_disponiveis[['pc']], headers=['PCs Disponíveis'], tablefmt="grid", showindex=True))

            # Selecionar horário
            horario_escolhido = input("\nDigite o horário que deseja agendar (ex: 08:00 - 09:00): ").strip()
            
            if horario_escolhido not in horarios_agrupados:
                msg("Horário inválido.", "err")
                continue

            pcs_horario = horarios_agrupados[horario_escolhido]
            pcs_disponiveis = pcs_horario[pcs_horario['status'] == 'Disponível']
            
            if pcs_disponiveis.empty:
                msg("Não há PCs disponíveis neste horário.", "warn")
                continue

            print(f"\nPCs disponíveis para {horario_escolhido}:")
            print(tabulate(pcs_disponiveis[['pc']], headers=['PC', 'Índice'], tablefmt="grid", showindex=True))
            
            # Selecionar múltiplos PCs
            indices_input = input("\nDigite os índices dos PCs que deseja agendar (separados por vírgula): ").strip()
            try:
                indices = [int(idx.strip()) for idx in indices_input.split(',')]
            except ValueError:
                msg("Formato inválido. Use números separados por vírgula.", "err")
                continue

            # Agendar múltiplos PCs
            resultados = agendamento_service.agendar_multiplos_pcs(indices, USERS[usuario_logado]["nome"])
            
            if resultados['sucessos']:
                criar_backup()
                msg("\n✅ Agendamentos realizados com sucesso:", "ok")
                for sucesso in resultados['sucessos']:
                    print(f"  ✓ {sucesso}")
            
            if resultados['falhas']:
                msg("\n❌ Alguns agendamentos falharam:", "err")
                for falha in resultados['falhas']:
                    print(f"  ✗ {falha}")

        elif escolha == "5":
            return
        else:
            msg("Opção inválida", "warn")

def limpar_dados(usuario_logado: str):
    if usuario_logado != "admin":
        msg("Apenas o ADMIN pode acessar esta opção.", "warn")
        return
    print("\n=== MENU DE LIMPEZA DE DADOS ===")
    print("1 - Limpar relatórios")
    print("2 - Limpar agendamentos")
    print("3 - Limpar alunos")
    print("4 - Limpar tudo")
    print("5 - Voltar")
    escolha = input("Escolha uma opção: ").strip()
    if escolha == "1":
        msg("Limpando relatórios...", "info")
        time.sleep(0.6)
        for arq in (ARQ_REL, ARQ_REL_XLSX):
            try:
                if arq.exists(): arq.unlink()
            except Exception:
                pass
        criar_backup()
        msg("Relatórios apagados com sucesso!", "ok")
    elif escolha == "2":
        msg("Limpando agendamentos...", "info")
        time.sleep(0.6)
        try:
            if ARQ_AG.exists(): ARQ_AG.unlink()
            xlsx = Path("agendamentos.xlsx")
            if xlsx.exists(): xlsx.unlink()
        except Exception:
            pass
        criar_backup()
        msg("Agendamentos apagados com sucesso!", "ok")
    elif escolha == "3":
        msg("Limpando alunos...", "info")
        time.sleep(0.6)
        for arq in (ARQ_ALUNOS, ARQ_ALUNOS_XLSX):
            try:
                if arq.exists(): arq.unlink()
            except Exception:
                pass
        criar_backup()
        msg("Alunos apagados com sucesso!", "ok")
    elif escolha == "4":
        msg("Limpando todos os dados do sistema...", "info")
        time.sleep(0.6)
        arquivos = [ARQ_REL, ARQ_REL_XLSX, ARQ_AG, ARQ_ALUNOS, ARQ_ALUNOS_XLSX, Path("agendamentos.xlsx")]
        for arq in arquivos:
            try:
                if arq.exists(): arq.unlink()
            except Exception:
                pass
        criar_backup()
        msg("Todos os dados foram apagados com sucesso!", "ok")
    elif escolha == "5":
        return
    else:
        msg("Opção inválida.", "warn")

def login():
    acesso_liberado = False
    usuario_logado = None
    while not acesso_liberado:
        print("\n=== LOGIN ===")
        usuario = input("Digite seu login: ").lower().strip()
        senha = pwinput.pwinput(prompt="Digite a senha: ", mask="*").strip()
        limpar_tela()
        if usuario in USERS and USERS[usuario]["senha"] == senha:
            msg("Login efetuado com sucesso", "ok")
            usuario_logado = usuario
            acesso_liberado = True
            time.sleep(0.4)
        else:
            msg("Login inválido. Tente novamente", "err")
            time.sleep(0.6)
            limpar_tela()
    return usuario_logado

def menu_principal(usuario: str):
    """Menu principal com interface melhorada"""
    opcoes = [
        ("1", "Computadores", menu_computadores),
        ("2", "Agendamento", menu_agendamento),
        ("3", "Relatório", gerar_relatorio),
        ("4", "Sair", lambda u: exit())
    ]
    
    if usuario == "admin":
        opcoes.append(("5", "Limpar dados", limpar_dados))
    
    while True:
        print(f"\n{'='*25}")
        print(f"=== MENU PRINCIPAL ===")
        print(f"Usuário: {USERS[usuario]['nome']}")
        print(f"{'='*25}")
        
        for codigo, descricao, _ in opcoes:
            print(f"{codigo} - {descricao}")
        
        escolha = input("\nEscolha uma opção: ").strip()
        
        for codigo, _, funcao in opcoes:
            if escolha == codigo:
                funcao(usuario)
                break
        else:
            msg("Opção inválida.", "warn")

def main():
    usuario = login()
    menu_principal(usuario)

if __name__ == "__main__":
    main()